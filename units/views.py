from django.shortcuts import render, redirect, get_object_or_404
from .models import MobileUnit, MedicalStaff, SupportStaff, Equipment, Intervention, Patient, Notification, Unit, MedicalRecord, UnitLocation
from .forms import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib.admin.views.decorators import staff_member_required
from django.forms import modelform_factory
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta, datetime
from django.db.models.functions import TruncMonth
from django.views.decorators.http import require_http_methods
import json
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from openpyxl.utils import get_column_letter

# Dashboard view
@login_required
def dashboard(request):
    if request.user.is_staff:
        # Admin Dashboard
        # Get total counts
        total_units = MobileUnit.objects.count()
        total_medical_staff = MedicalStaff.objects.count()
        total_support_staff = SupportStaff.objects.count()
        total_patients = Patient.objects.count()

        # Get interventions by unit
        interventions_by_unit = Intervention.objects.values('unit__name').annotate(count=Count('id')).order_by('-count')
        unit_names = [item['unit__name'] for item in interventions_by_unit]
        interventions_by_unit = [item['count'] for item in interventions_by_unit]

        # Get patients treated over time (last 6 months)
        six_months_ago = timezone.now() - timedelta(days=180)
        patients_over_time = MedicalRecord.objects.filter(
            date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')

        months = [item['month'].strftime('%B %Y') for item in patients_over_time]
        patients_over_time = [item['count'] for item in patients_over_time]

        # Get staff distribution
        medical_staff_count = MedicalStaff.objects.filter(role='doctor').count()
        nursing_staff_count = MedicalStaff.objects.filter(role='nurse').count()
        support_staff_count = SupportStaff.objects.count()

        # Get equipment counts by unit
        equipment_by_unit = Equipment.objects.values('unit__name').annotate(count=Count('id')).order_by('-count')
        equipment_unit_names = [item['unit__name'] for item in equipment_by_unit]
        equipment_counts = [item['count'] for item in equipment_by_unit]

        context = {
            'total_units': total_units,
            'total_medical_staff': total_medical_staff,
            'total_support_staff': total_support_staff,
            'total_patients': total_patients,
            'unit_names': unit_names,
            'interventions_by_unit': interventions_by_unit,
            'months': months,
            'patients_over_time': patients_over_time,
            'medical_staff_count': medical_staff_count,
            'nursing_staff_count': nursing_staff_count,
            'support_staff_count': support_staff_count,
            'equipment_unit_names': equipment_unit_names,
            'equipment_counts': equipment_counts,
        }
    elif hasattr(request.user, 'medicalstaff'):
        # Medical staff dashboard
        medical_staff = request.user.medicalstaff
        assigned_unit = medical_staff.assigned_unit
        
        if not assigned_unit:
            messages.warning(request, "Vous n'êtes pas assigné à une unité mobile.")
            return redirect('profile')
        
        # Get counts for the assigned unit
        patients_count = Patient.objects.filter(medical_records__unit=assigned_unit).distinct().count()
        interventions_count = Intervention.objects.filter(unit=assigned_unit).count()
        equipment_count = Equipment.objects.filter(unit=assigned_unit).count()
        medical_records_count = MedicalRecord.objects.filter(unit=assigned_unit).count()
        
        # Get recent medical records for the assigned unit
        recent_medical_records = MedicalRecord.objects.filter(
            unit=assigned_unit
        ).select_related('patient', 'doctor').order_by('-date')[:5]
        
        context = {
            'patients_count': patients_count,
            'interventions_count': interventions_count,
            'equipment_count': equipment_count,
            'medical_records_count': medical_records_count,
            'recent_medical_records': recent_medical_records,
        }
    elif hasattr(request.user, 'supportstaff'):
        # Support staff dashboard
        support_staff = request.user.supportstaff
        assigned_unit = support_staff.assigned_unit
        
        if not assigned_unit:
            messages.warning(request, "Vous n'êtes pas assigné à une unité mobile.")
            return redirect('profile')
        
        # Get counts for the assigned unit
        equipment_count = Equipment.objects.filter(unit=assigned_unit).count()
        interventions_count = Intervention.objects.filter(unit=assigned_unit).count()
        
        context = {
            'equipment_count': equipment_count,
            'interventions_count': interventions_count,
        }
    else:
        messages.error(request, "Vous n'avez pas les permissions nécessaires pour accéder au tableau de bord.")
        return redirect('login')

    return render(request, 'units/dashboard.html', context)

def generic_edit(request, model_class, form_class, redirect_url, obj_id=None):
    ModelForm = modelform_factory(model_class, form=form_class)
    if obj_id:
        obj = get_object_or_404(model_class, id=obj_id)
        form = ModelForm(request.POST or None, instance=obj)
    else:
        form = ModelForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect(redirect_url)

    return render(request, 'units/generic_form.html', {'form': form})

# Unit views
@login_required
def unit_list(request):
    units = MobileUnit.objects.all()
    
    # Get filter parameters
    status = request.GET.get('status')
    search_query = request.GET.get('search')
    sort = request.GET.get('sort')
    
    # Apply filters
    if status:
        units = units.filter(status=status)
    
    if search_query:
        units = units.filter(
            Q(name__icontains=search_query) |
            Q(location__icontains=search_query)
        )
    
    # Apply sorting
    if sort:
        units = units.order_by(sort)
    else:
        units = units.order_by('-start_date')  # Default sort by start date descending
    
    context = {
        'units': units,
        'current_status': status,
        'search_query': search_query,
        'current_sort': sort,
    }
    
    return render(request, 'units/unit_list.html', context)

@login_required
def unit_create(request):
    return generic_edit(request, MobileUnit, MobileUnitForm, 'unit_list')

@login_required
def unit_update(request, unit_id):
    return generic_edit(request, MobileUnit, MobileUnitForm, 'unit_list', unit_id)

@login_required
def unit_delete(request, unit_id):
    unit = get_object_or_404(MobileUnit, id=unit_id)
    if request.method == 'POST':
        unit.delete()
        return redirect('unit_list')
    return render(request, 'units/confirm_delete.html', {'object': unit})

# Staff views
@login_required
def staff_list(request):
    staff = MedicalStaff.objects.all()
    
    # Get filter parameters
    search_query = request.GET.get('search')
    sort = request.GET.get('sort')
    
    # Apply filters
    if search_query:
        staff = staff.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Apply sorting
    if sort:
        staff = staff.order_by(sort)
    else:
        staff = staff.order_by('name')  # Default sort by name
    
    context = {
        'staff_list': staff,
        'search_query': search_query,
        'current_sort': sort,
    }
    
    return render(request, 'units/staff_list.html', context)

@login_required
def staff_detail(request, staff_id):
    staff = get_object_or_404(MedicalStaff, id=staff_id)
    return render(request, 'units/staff_detail.html', {'staff': staff})

@login_required
def staff_create(request):
    if request.method == 'POST':
        form = MedicalStaffForm(request.POST)
        if form.is_valid():
            # Get the part before @ in email for username
            email = form.cleaned_data['email']
            username = email.split('@')[0]
            
            # Create a new user account
            user = User.objects.create_user(
                username=username,
                email=email,
                password='20252025',  # Set default password
                first_name=form.cleaned_data['name'].split()[0] if ' ' in form.cleaned_data['name'] else form.cleaned_data['name'],
                last_name=form.cleaned_data['name'].split()[-1] if ' ' in form.cleaned_data['name'] else ''
            )
            
            # Create the medical staff member
            staff = form.save(commit=False)
            staff.user = user
            staff.save()
            
            messages.success(request, f'Le membre du personnel a été créé avec succès. Identifiants de connexion: Username: {username}, Mot de passe: 20252025')
            return redirect('staff_list')
    else:
        form = MedicalStaffForm()
    
    return render(request, 'units/generic_form.html', {'form': form})

@login_required
def staff_update(request, staff_id):
    return generic_edit(request, MedicalStaff, MedicalStaffForm, 'staff_list', staff_id)

@login_required
def staff_delete(request, staff_id):
    staff = get_object_or_404(MedicalStaff, id=staff_id)
    if request.method == 'POST':
        staff.delete()
        return redirect('staff_list')
    return render(request, 'units/confirm_delete.html', {'object': staff})

# Support views
@login_required
def support_list(request):
    support = SupportStaff.objects.all()
    
    # Get filter parameters
    search_query = request.GET.get('search')
    sort = request.GET.get('sort')
    
    # Apply filters
    if search_query:
        support = support.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Apply sorting
    if sort:
        support = support.order_by(sort)
    else:
        support = support.order_by('name')  # Default sort by name
    
    context = {
        'support': support,
        'search_query': search_query,
        'current_sort': sort,
    }
    
    return render(request, 'units/support_list.html', context)

@login_required
def support_create(request):
    return generic_edit(request, SupportStaff, SupportStaffForm, 'support_list')

@login_required
def support_update(request, support_id):
    return generic_edit(request, SupportStaff, SupportStaffForm, 'support_list', support_id)

@login_required
def support_delete(request, support_id):
    support = get_object_or_404(SupportStaff, id=support_id)
    if request.method == 'POST':
        support.delete()
        return redirect('support_list')
    return render(request, 'units/confirm_delete.html', {'object': support})

# Equipment views
@login_required
def equipment_list(request):
    equipment = Equipment.objects.all()
    
    # Get filter parameters
    status = request.GET.get('status')
    search_query = request.GET.get('search')
    sort = request.GET.get('sort')
    
    # Apply filters
    if status:
        equipment = equipment.filter(status=status)
    
    if search_query:
        equipment = equipment.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Apply sorting
    if sort:
        equipment = equipment.order_by(sort)
    else:
        equipment = equipment.order_by('name')  # Default sort by name
    
    context = {
        'equipment_list': equipment,
        'current_status': status,
        'search_query': search_query,
        'current_sort': sort,
    }
    
    return render(request, 'units/equipment_list.html', context)

@login_required
def equipment_create(request):
    return generic_edit(request, Equipment, EquipmentForm, 'equipment_list')

@login_required
def equipment_update(request, equipment_id):
    return generic_edit(request, Equipment, EquipmentForm, 'equipment_list', equipment_id)

@login_required
def equipment_delete(request, equipment_id):
    equipment = get_object_or_404(Equipment, id=equipment_id)
    if request.method == 'POST':
        equipment.delete()
        return redirect('equipment_list')
    return render(request, 'units/confirm_delete.html', {'object': equipment})

@login_required
def equipment_detail(request, equipment_id):
    equipment = get_object_or_404(Equipment, id=equipment_id)
    return render(request, 'units/equipment_detail.html', {'equipment': equipment})

# Intervention views
@login_required
def intervention_list(request):
    # Get filter parameters
    status = request.GET.get('status')
    search_query = request.GET.get('search')
    sort = request.GET.get('sort')
    
    # Start with base queryset
    interventions = Intervention.objects.all()
    
    # Check if user is medical staff
    try:
        medical_staff = MedicalStaff.objects.get(email=request.user.email)
        if medical_staff.assigned_unit:
            # Filter interventions for medical staff's unit
            interventions = interventions.filter(unit=medical_staff.assigned_unit)
    except MedicalStaff.DoesNotExist:
        pass
    
    # Apply filters
    if status:
        interventions = interventions.filter(status=status)
    
    if search_query:
        interventions = interventions.filter(
            Q(location__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Apply sorting
    if sort:
        interventions = interventions.order_by(sort)
    else:
        interventions = interventions.order_by('-date')  # Default sort by date descending
    
    context = {
        'interventions': interventions,
        'current_status': status,
        'search_query': search_query,
        'current_sort': sort,
    }
    
    return render(request, 'units/intervention_list.html', context)

@login_required
def intervention_create(request):
    return generic_edit(request, Intervention, InterventionForm, 'intervention_list')

@login_required
def intervention_update(request, intervention_id):
    return generic_edit(request, Intervention, InterventionForm, 'intervention_list', intervention_id)

@login_required
def intervention_delete(request, intervention_id):
    intervention = get_object_or_404(Intervention, id=intervention_id)
    if request.method == 'POST':
        intervention.delete()
        return redirect('intervention_list')
    return render(request, 'units/confirm_delete.html', {'object': intervention})

@login_required
def intervention_detail(request, intervention_id):
    intervention = get_object_or_404(Intervention, id=intervention_id)
    return render(request, 'units/intervention_detail.html', {'intervention': intervention})

# Patient views
@login_required
def patient_list(request):
    patients = Patient.objects.all().order_by('-created_at')
    search_query = request.GET.get('search', '')
    if search_query:
        patients = patients.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    return render(request, 'units/patient_list.html', {
        'patients': patients,
        'search_query': search_query
    })

@login_required
def patient_detail(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    medical_records = patient.medical_records.all().order_by('-date')
    return render(request, 'units/patient_detail.html', {
        'patient': patient,
        'medical_records': medical_records
    })

@login_required
def patient_create(request):
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save()
            return redirect('patient_detail', patient_id=patient.id)
    else:
        form = PatientForm()
    return render(request, 'units/patient_form.html', {'form': form})

@login_required
def patient_update(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            return redirect('patient_detail', patient_id=patient.id)
    else:
        form = PatientForm(instance=patient)
    return render(request, 'units/patient_form.html', {'form': form, 'patient': patient})

@login_required
def medical_record_create(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    if request.method == 'POST':
        form = MedicalRecordForm(request.POST, user=request.user)
        if form.is_valid():
            medical_record = form.save(commit=False)
            medical_record.patient = patient
            medical_record.save()
            messages.success(request, 'Le dossier médical a été créé avec succès.')
            return redirect('patient_detail', patient_id=patient.id)
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = MedicalRecordForm(user=request.user)
    
    return render(request, 'units/medical_record_form.html', {
        'form': form,
        'patient': patient,
        'form_errors': form.errors if request.method == 'POST' else None
    })

@login_required
def medical_record_update(request, record_id):
    medical_record = get_object_or_404(MedicalRecord, id=record_id)
    if request.method == 'POST':
        form = MedicalRecordForm(request.POST, instance=medical_record)
        if form.is_valid():
            form.save()
            return redirect('patient_detail', patient_id=medical_record.patient.id)
    else:
        form = MedicalRecordForm(instance=medical_record)
    return render(request, 'units/medical_record_form.html', {
        'form': form,
        'patient': medical_record.patient,
        'record': medical_record
    })

@login_required
def medical_record_detail(request, record_id):
    medical_record = get_object_or_404(MedicalRecord, id=record_id)
    return render(request, 'units/medical_record_detail.html', {'record': medical_record})

@login_required
def patient_delete(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    if request.method == 'POST':
        patient.delete()
        return redirect('patient_list')
    return render(request, 'units/confirm_delete.html', {'object': patient})

# Auth views
def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard')
    else:
        form = SignUpForm()
    return render(request, 'units/signup.html', {'form': form})

@login_required
def profile(request):
    return render(request, 'units/profile.html')

@staff_member_required
def user_list(request):
    users = User.objects.all()
    return render(request, 'units/user_list.html', {'users': users})

@login_required
def unit_calendar(request):
    units = MobileUnit.objects.all()
    return render(request, 'units/unit_calendar.html', {'units': units})

@login_required
def unit_detail(request, pk):
    unit = get_object_or_404(MobileUnit, pk=pk)
    return render(request, 'units/unit_detail.html', {'unit': unit})

@login_required
def mark_notification_read(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id)
    notification.is_read = True
    notification.save()
    return JsonResponse({'status': 'success'})

@login_required
def unit_tracking(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    locations = UnitLocation.objects.filter(unit=unit).order_by('-timestamp')[:100]
    return render(request, 'units/unit_tracking.html', {
        'unit': unit,
        'locations': locations
    })

@login_required
@require_http_methods(["POST"])
def update_unit_location(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    try:
        data = json.loads(request.body)
        location = UnitLocation.objects.create(
            unit=unit,
            latitude=data['latitude'],
            longitude=data['longitude'],
            speed=data.get('speed'),
            heading=data.get('heading'),
            battery_level=data.get('battery_level'),
            signal_strength=data.get('signal_strength')
        )
        unit.current_latitude = data['latitude']
        unit.current_longitude = data['longitude']
        unit.last_location_update = timezone.now()
        unit.battery_level = data.get('battery_level')
        unit.signal_strength = data.get('signal_strength')
        unit.is_online = True
        unit.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@login_required
def reports_dashboard(request):
    if not request.user.is_staff:
        messages.error(request, "Vous n'avez pas la permission d'accéder à cette page.")
        return redirect('dashboard')
    
    units = MobileUnit.objects.all()
    context = {
        'units': units,
    }
    return render(request, 'units/reports_dashboard.html', context)

def get_unit_doctors(request, unit_id):
    """API endpoint to get doctors assigned to a specific unit."""
    try:
        unit = Unit.objects.get(id=unit_id)
        doctors = User.objects.filter(
            medicalstaff__assigned_unit=unit
        ).values('id', 'first_name', 'last_name')
        return JsonResponse(list(doctors), safe=False)
    except Unit.DoesNotExist:
        return JsonResponse({'error': 'Unit not found'}, status=404)

@login_required
def profile_update(request):
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Votre profil a été mis à jour avec succès.')
            return redirect('profile')
    else:
        form = UserUpdateForm(instance=request.user)
    
    return render(request, 'units/profile_update.html', {
        'form': form
    })

def generate_unit_report(request, unit_id):
    unit = get_object_or_404(MobileUnit, id=unit_id)
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Define styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add title
    ws['A1'] = f"Rapport de l'unité {unit.name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws.merge_cells('A1:E1')
    
    # Add date
    ws['A2'] = f"Date du rapport : {timezone.now().strftime('%d/%m/%Y')}"
    ws['A2'].alignment = center_alignment
    ws.merge_cells('A2:E2')
    
    # Unit Information
    current_row = 4
    ws[f'A{current_row}'] = "Informations de l'unité"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    
    unit_info = [
        ["Nom", unit.name],
        ["Localisation", unit.location],
        ["Statut", unit.status],
        ["Date de début", unit.start_date.strftime('%d/%m/%Y')],
        ["Date de fin", unit.end_date.strftime('%d/%m/%Y') if unit.end_date else "Non définie"]
    ]
    
    for row_idx, (label, value) in enumerate(unit_info, start=current_row + 1):
        ws[f'A{row_idx}'] = label
        ws[f'B{row_idx}'] = value
    
    # Medical Staff Information
    current_row = len(unit_info) + current_row + 2
    ws[f'A{current_row}'] = "Personnel médical"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    
    staff_headers = ["Nom", "Rôle", "Email", "Téléphone", "Date d'embauche"]
    for col_idx, header in enumerate(staff_headers, start=1):
        cell = ws.cell(row=current_row + 1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    medical_staff = MedicalStaff.objects.filter(assigned_unit=unit)
    for row_idx, staff in enumerate(medical_staff, start=current_row + 2):
        ws[f'A{row_idx}'] = staff.name
        ws[f'B{row_idx}'] = staff.get_role_display()
        ws[f'C{row_idx}'] = staff.email
        ws[f'D{row_idx}'] = staff.phone_number or "Non renseigné"
    
    # Equipment Information
    current_row = current_row + len(medical_staff) + 3
    ws[f'A{current_row}'] = "Équipements"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    
    equipment_headers = ["Nom", "Quantité", "Statut"]
    for col_idx, header in enumerate(equipment_headers, start=1):
        cell = ws.cell(row=current_row + 1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    equipment = Equipment.objects.filter(unit=unit)
    for row_idx, item in enumerate(equipment, start=current_row + 2):
        ws[f'A{row_idx}'] = item.name
        ws[f'B{row_idx}'] = item.quantity
        ws[f'C{row_idx}'] = "En stock" if item.quantity > 0 else "Rupture de stock"
    
    # Interventions Information
    current_row = current_row + len(equipment) + 3
    ws[f'A{current_row}'] = "Interventions"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    
    intervention_headers = ["Date", "Localisation", "Description", "Patients traités"]
    for col_idx, header in enumerate(intervention_headers, start=1):
        cell = ws.cell(row=current_row + 1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    interventions = Intervention.objects.filter(unit=unit).order_by('-date')
    for row_idx, intervention in enumerate(interventions, start=current_row + 2):
        ws[f'A{row_idx}'] = intervention.date.strftime('%d/%m/%Y')
        ws[f'B{row_idx}'] = intervention.location
        ws[f'C{row_idx}'] = intervention.description
        ws[f'D{row_idx}'] = intervention.patients_treated
    
    # Adjust column widths
    for col in range(1, 6):  # A to E
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=rapport_unite_{unit.name}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response